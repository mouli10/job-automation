import pandas as pd
from datetime import datetime
from src.config import DATA_DIR, MAX_REPORT_JOBS
from src.db.database import SessionLocal
from src.db.models import Job, Score, Resume, Report


def generate_daily_report(app_assist_data: dict = None) -> str:
    """
    Generates the daily Excel report with top MAX_REPORT_JOBS jobs sorted by ATS score.
    Columns: Role Name, Company, Location, Job Link, Posting Date,
             Suggested Resume, Resume Type, ATS Score, Google Drive Link
    """
    db = SessionLocal()
    try:
        today = datetime.utcnow().date()

        # Get all scores from today
        all_scores = db.query(Score).all()
        today_scores = [s for s in all_scores if s.created_at and s.created_at.date() == today]

        # Group all scores by job to aggregate them
        job_scores: dict = {}
        for s in today_scores:
            if s.ats_score is None:
                continue
            if s.job_id not in job_scores:
                job_scores[s.job_id] = []
            job_scores[s.job_id].append(s)

        data = []
        for job_id, score_list in job_scores.items():
            job = db.query(Job).filter(Job.id == job_id).first()
            if not job:
                continue

            # Identify the best score record for this job
            best_score_rec = max(score_list, key=lambda x: x.ats_score)
            
            from src.config import admin_config
            min_ats = admin_config["optional_filters"].get("min_ats_score", 0.0)
            if (best_score_rec.ats_score or 0.0) < min_ats:
                continue
                
            best_resume = db.query(Resume).filter(Resume.id == best_score_rec.resume_id).first()
            if not best_resume:
                continue

            # Format all scores into a single string: "Mouli_V_T.docx-8.0, Mouli_V_B.docx-7.5"
            all_scores_str = []
            for s in score_list:
                r = db.query(Resume).filter(Resume.id == s.resume_id).first()
                if r:
                    all_scores_str.append(f"{r.filename}-{round(s.ats_score, 1)}")

            # Process missing keywords
            missing_kw_str = best_score_rec.missing_keywords or "None"
            review_str = best_score_rec.review or "No AI Review"
            review_text = f"MISSING:\n{missing_kw_str}\n\nREVIEW:\n{review_str}"
            # Calculate final sorted score with optional Admin Priority Boosts
            final_score = best_score_rec.ats_score if best_score_rec.ats_score else 0.0
            
            # 1. Company Priority
            priority_companies = admin_config["optional_filters"].get("priority_companies", [])
            if any(pc.lower() in (job.company or "").lower() for pc in priority_companies):
                final_score += 2.0
                
            # 2. Keyword Priority
            if admin_config["optional_filters"].get("keyword_priority_boost", False):
                priority_keywords = admin_config["optional_filters"].get("priority_keywords", [])
                if any(pk.lower() in (job.title or "").lower() for pk in priority_keywords):
                    final_score += 2.0

            # Format the dictionary
            data.append({
                "Role Name": job.title or "",
                "Company": job.company or "",
                "Location": job.location or "",
                "Job Link": job.link or "",
                "Posting Date": job.posting_date or "Recent (24h)",
                "All Resume Scores": ", ".join(all_scores_str),
                "Suggested Resume": f"[OPT] {best_score_rec.gdrive_link}" if best_score_rec.gdrive_link else best_resume.filename or "",
                "Fit Analysis Summary": app_assist_data.get(job.id, {}).get("why_fit", "") if app_assist_data else "",
                "Cover Letter": app_assist_data.get(job.id, {}).get("cover_letter_path", "") if app_assist_data else "",
                "AI Review & Missing Keywords": review_text,
                "_best_score": final_score,
            })

        # Sort data descending by the Best Score that we saved in a temporary key
        data.sort(key=lambda x: x["_best_score"], reverse=True)
        data = data[:MAX_REPORT_JOBS]

        # Remove the temporary sort key before making the DataFrame
        for d in data:
            del d["_best_score"]

        df = pd.DataFrame(data)

        today_str = today.strftime('%Y%m%d')
        report_filename = f"Daily_Job_Matches_{today_str}.xlsx"
        from src.config import REPORTS_DIR
        report_path = REPORTS_DIR / report_filename

        df.to_excel(report_path, index=False)

        # Format the Excel file for readability
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = openpyxl.load_workbook(report_path)
            ws = wb.active

            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)

            col_widths = {1: 28, 2: 22, 3: 15, 4: 45, 5: 15, 6: 35, 7: 25, 8: 60, 9: 30, 10: 60}
            for col_idx, cell in enumerate(ws[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_idx, 18)

            # Alternate row shading for readability
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                fill_color = "D6E4F0" if row_idx % 2 == 0 else "FFFFFF"
                row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                for cell in row:
                    cell.fill = row_fill

            ws.freeze_panes = "A2"  # Freeze header row
            wb.save(report_path)
        except Exception:
            pass  # Formatting is optional — file is still valid without it

        # Record in DB
        report_record = Report(filepath=str(report_path), jobs_processed=len(df))
        db.add(report_record)
        db.commit()

        return str(report_path)
    finally:
        db.close()
